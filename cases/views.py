from .forms import CaseDistributionForm # Importamos aquí para evitar errores circulares
import json
import logging
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied, ValidationError as DjangoValidationError
from django.db import transaction
from django.db.models import Count, Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.views import View
from django.views.generic import DetailView, ListView, TemplateView
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .forms import CaseDistributionForm
from accounts.models import User
from accounts.services import notify_advisor_document_uploaded, notify_case_participants_bitacora_entry
from accounts.views import get_type_label
from .forms import BitacoraEntryForm, CaseDistributionForm
from .bitacora_summary import build_case_bitacora_summary
from .models import (
    AdvisorAvailabilitySlot, BitacoraDocument, BitacoraEntry,
    Case, CaseAppointment, CaseAssignment, CaseCalendarEvent, CaseDeadline, Notification,
)
from .services import (
    CaseCompletionValidator,
    CaseDistributionService,
    CaseDocumentService,
    CaseStageManager,
    DelayedCasesDashboardService,
    MissedAppointmentsDashboardService,
    asignar_caso_automatico,
    CaseAutomaticDistributionService,
    get_category_type_choices,
    UNCATEGORIZED_CATEGORY_KEY,
    UNTYPED_CASE_TYPE_KEY,
)
from .teams import create_teams_calendar_event


def get_topbar_user_data(user):
    return {
        "full_name": user.full_name,
        "role_name": user.get_role_display(),
        "initials": user.initials,
    }


class SecretaryCasesView(LoginRequiredMixin, TemplateView):
    template_name = "cases/secretary_cases.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != User.Role.SECRETARIA:
            messages.error(request, "No autorizado.")
            return redirect("dashboard_redirect")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cases = Case.objects.all().order_by("-created_at")

        # Categorías (Salas)
        categories = [
            ("PEN", "Sala Penal"),
            ("LAB", "Sala Laboral"),
            ("CIV", "Sala Civil"),
            ("FAM", "Sala Familia"),
            ("DER_PUB_MIG", "Sala Derecho Público"),
            ("PUB", "Sala derecho público"),
            ("MIGR", "Sala derecho público - Migrantes"),
        ]

        # Tipos (Trámites) completos por sala
        type_options = [
            # Sala Penal
            {"key": "PROC",     "label": "Proceso",               "category": "PEN"},
            {"key": "DER_PET",  "label": "Derecho de petición",   "category": "PEN"},
            {"key": "TUT",      "label": "Tutela",                "category": "PEN"},
            {"key": "CONC_DEN", "label": "Concepto + denuncia",   "category": "PEN"},
            {"key": "CONC",     "label": "Concepto",              "category": "PEN"},
            {"key": "MEM",      "label": "Memorial",              "category": "PEN"},
            # Sala Laboral
            {"key": "PROC",     "label": "Proceso",               "category": "LAB"},
            {"key": "LIQ",      "label": "Liquidación",           "category": "LAB"},
            {"key": "LIQ_CONC", "label": "Liquidación + concepto","category": "LAB"},
            {"key": "TUT",      "label": "Tutela",                "category": "LAB"},
            {"key": "DER_PET",  "label": "Derecho de petición",   "category": "LAB"},
            {"key": "CONC",     "label": "Concepto",              "category": "LAB"},
            {"key": "QUE",      "label": "Queja",                 "category": "LAB"},
            {"key": "MEM",      "label": "Memorial",              "category": "LAB"},
            # Sala Civil
            {"key": "PROC",     "label": "Proceso",               "category": "CIV"},
            {"key": "COB_PRE",  "label": "Cobro pre-jurídico",    "category": "CIV"},
            {"key": "TUT",      "label": "Tutela",                "category": "CIV"},
            {"key": "DER_PET",  "label": "Derecho de petición",   "category": "CIV"},
            {"key": "CONC_DP",  "label": "Concepto + DP",         "category": "CIV"},
            {"key": "QUE",      "label": "Queja",                 "category": "CIV"},
            {"key": "MEM",      "label": "Memorial",              "category": "CIV"},
            {"key": "CONC",     "label": "Concepto",              "category": "CIV"},
            {"key": "CLI_EMP",  "label": "Clínica empresarial",   "category": "CIV"},
            # Sala Familia
            {"key": "PROC",     "label": "Proceso",               "category": "FAM"},
            {"key": "CONC_DP",  "label": "Concepto + DP",         "category": "FAM"},
            {"key": "DER_PET",  "label": "Derecho de petición",   "category": "FAM"},
            {"key": "TUT",      "label": "Tutela",                "category": "FAM"},
            {"key": "MEM",      "label": "Memorial",              "category": "FAM"},
            {"key": "QUE",      "label": "Queja",                 "category": "FAM"},
            {"key": "COB_PRE",  "label": "Cobro pre-jurídico",    "category": "FAM"},
            {"key": "CONC",     "label": "Concepto",              "category": "FAM"},
            # Sala Derecho Público
            {"key": "SOL_REF",     "label": "Solicitud de refugio",          "category": "DER_PUB_MIG"},
            {"key": "SOL_REF_DP",  "label": "Solicitud de refugio + DP",     "category": "DER_PUB_MIG"},
            {"key": "SOL_REF_TUT", "label": "Solicitud de refugio + Tutela", "category": "DER_PUB_MIG"},
            {"key": "TRAM_SAL",    "label": "Trámite salvoconducto",         "category": "DER_PUB_MIG"},
            {"key": "TUT",         "label": "Tutela",                        "category": "DER_PUB_MIG"},
            {"key": "CONC_DP",     "label": "Concepto + DP",                 "category": "DER_PUB_MIG"},
            {"key": "DER_PET",     "label": "Derecho de petición",           "category": "DER_PUB_MIG"},
            {"key": "CONC",        "label": "Concepto",                      "category": "DER_PUB_MIG"},
            # Sala Derecho público
            {"key": "PROC",        "label": "Proceso",                       "category": "PUB"},
            {"key": "CONC_DP",     "label": "Concepto + DP",                 "category": "PUB"},
            {"key": "DER_PET",     "label": "Derecho de petición",           "category": "PUB"},
            {"key": "TUT",         "label": "Tutela",                        "category": "PUB"},
            {"key": "MEM",         "label": "Memorial",                      "category": "PUB"},
            {"key": "QUE",         "label": "Queja",                         "category": "PUB"},
            {"key": "COB_PRE",     "label": "Cobro pre-jurídico",            "category": "PUB"},
            {"key": "CONC",        "label": "Concepto",                      "category": "PUB"},
            {"key": "MIG",         "label": "Migrantes",                     "category": "PUB"},
            # Sala Derecho público - Migrantes
            {"key": "SOL_REF",     "label": "Solicitud de refugio",          "category": "MIGR"},
            {"key": "SOL_REF_DP",  "label": "Solicitud de refugio + DP",     "category": "MIGR"},
            {"key": "SOL_REF_TUT", "label": "Solicitud de refugio + Tutela", "category": "MIGR"},
            {"key": "TRAM_SAL",    "label": "Trámite salvoconducto",         "category": "MIGR"},
            {"key": "TUT",         "label": "Tutela",                        "category": "MIGR"},
            {"key": "CONC_DP",     "label": "Concepto + DP",                 "category": "MIGR"},
            {"key": "DER_PET",     "label": "Derecho de petición",           "category": "MIGR"},
            {"key": "CONC",        "label": "Concepto",                      "category": "MIGR"},
        ]

        # Estados
        statuses = [
            ("SIN", "Sin asignar"),
            ("AUT", "Autoasignado"),
            ("ASSIGNED_FLOW", "Asignados / en proceso"),
        ]

        context["user_data"] = get_topbar_user_data(self.request.user)
        context["cases"] = cases
        context["cases_count"] = cases.count()
        context["categories"] = categories
        context["type_options"] = type_options
        context["statuses"] = statuses
        context["students"] = User.objects.filter(
            role=User.Role.ESTUDIANTE,
            is_active=True,
        ).order_by("first_name", "last_name")
        return context


class CaseReportView(LoginRequiredMixin, TemplateView):
    """
    Vista para mostrar un reporte de todos los casos del consultorio jurídico.
    Accesible para Secretaria y Asesor.
    Permite descargar los datos en Excel.
    """
    template_name = "cases/case_report.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ASESOR]:
            messages.error(request, "No tienes permiso para acceder a esta funcionalidad.")
            return redirect("dashboard_redirect")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cases = Case.objects.select_related(
            "beneficiary",
            "assigned_student",
            "advisor",
            "secretary"
        ).all().order_by("-created_at")

        context["user_data"] = get_topbar_user_data(self.request.user)
        context["cases"] = cases
        context["cases_count"] = cases.count()
        beneficiaries = User.objects.filter(role=User.Role.BENEFICIARIO)
        total_bens = beneficiaries.count()
        strata_counts = [
            {"label": f"Estrato {s}", "count": beneficiaries.filter(socioeconomic_stratum=s).count()}
            for s in range(1, 7)
        ]
        strata_max = max((item["count"] for item in strata_counts), default=1) or 1
        for item in strata_counts:
            item["pct"] = round(item["count"] / strata_max * 100)
        context["beneficiary_stats"] = {
            "total": total_bens,
            "with_gender": beneficiaries.exclude(gender__isnull=True).exclude(gender="").count(),
            "with_stratum": beneficiaries.exclude(socioeconomic_stratum__isnull=True).count(),
            "male_count": beneficiaries.filter(gender=User.Gender.MASCULINO).count(),
            "female_count": beneficiaries.filter(gender=User.Gender.FEMENINO).count(),
            "gender": [
                {"label": label, "count": beneficiaries.filter(gender=value).count()}
                for value, label in User.Gender.choices
            ],
            "strata": strata_counts,
        }
        return context

    def post(self, request, *args, **kwargs):
        """Descarga los casos en Excel"""
        if request.POST.get("download_excel") == "true":
            return self.download_excel()
        return self.render_to_response(self.get_context_data())

    def download_excel(self):
        """Genera y descarga archivo Excel con todos los casos"""
        cases = Case.objects.select_related(
            "beneficiary",
            "assigned_student",
            "advisor",
            "secretary"
        ).all().order_by("-created_at")

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Casos"

        # Estilos
        header_fill = PatternFill(start_color="5B5CE2", end_color="5B5CE2", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            "ID",
            "Beneficiario",
            "Documento beneficiario",
            "Sexo beneficiario",
            "Estrato beneficiario",
            "Correo beneficiario",
            "Estudiante Asignado",
            "Asesor",
            "Sala",
            "Trámite jurídico",
            "Fecha Creación",
            "Estado",
            "Etapa Actual"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # Datos
        for row, case in enumerate(cases, 2):
            ws.cell(row=row, column=1, value=case.case_number)
            ws.cell(row=row, column=2, value=case.beneficiary.full_name if case.beneficiary else "N/A")
            ws.cell(row=row, column=3, value=case.beneficiary.document_number if case.beneficiary else "N/A")
            ws.cell(row=row, column=4, value=case.beneficiary.get_gender_display() if case.beneficiary and case.beneficiary.gender else "No registrado")
            ws.cell(row=row, column=5, value=case.beneficiary.socioeconomic_stratum if case.beneficiary and case.beneficiary.socioeconomic_stratum else "No registrado")
            ws.cell(row=row, column=6, value=case.beneficiary.email if case.beneficiary and case.beneficiary.email else "No registrado")
            ws.cell(row=row, column=7, value=case.assigned_student.full_name if case.assigned_student else "Sin asignar")
            ws.cell(row=row, column=8, value=case.advisor.full_name if case.advisor else "N/A")
            ws.cell(row=row, column=9, value=case.get_category_display() if case.category else "N/A")
            ws.cell(row=row, column=10, value=case.get_specific_type_display() or "N/A")
            ws.cell(row=row, column=11, value=case.created_at.strftime("%d/%m/%Y") if case.created_at else "N/A")
            ws.cell(row=row, column=12, value=case.get_status_display())
            ws.cell(row=row, column=13, value=case.get_current_stage_display() if case.current_stage else "N/A")

        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = 18

        # Respuesta
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="reporte_casos.xlsx"'
        wb.save(response)
        return response


class CaseDetailView(LoginRequiredMixin, TemplateView):
    template_name = "cases/case_detail.html"

    def dispatch(self, request, *args, **kwargs):
        self.case = get_object_or_404(
            Case.objects.select_related(
                "beneficiary",
                "assigned_student",
                "advisor",
                "secretary",
            ),
            id=self.kwargs["case_id"],
        )
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        history_items = (
            BitacoraEntry.objects.filter(case=self.case)
            .select_related("author")
            .order_by("-created_at")
        )
        documents_count = self.case.case_documents.count()
        stages = CaseStageManager.get_stage_display_info(self.case)

        context["user_data"] = get_topbar_user_data(self.request.user)
        context["case"] = self.case
        context["can_reassign"] = self.request.user.role == User.Role.SECRETARIA
        context["can_change_stage"] = self.request.user.role in [
            User.Role.SECRETARIA,
            User.Role.ASESOR,
        ]
        context["bitacora_count"] = history_items.count()
        context["documents_count"] = documents_count
        context["history_items"] = history_items[:5]
        context["stages"] = stages
        context["closure_types"] = Case.ClosureType.choices
        return context


class DelayedCasesDashboardView(LoginRequiredMixin, TemplateView):
    template_name = "cases/delayed_cases_dashboard.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ASESOR, User.Role.ESTUDIANTE]:
            messages.error(request, "No autorizado.")
            return redirect("dashboard_redirect")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dashboard_data = DelayedCasesDashboardService.get_dashboard_data(self.request.user)

        role = self.request.user.role
        if role == User.Role.SECRETARIA:
            return_url = "secretary_dashboard"
        elif role == User.Role.ASESOR:
            return_url = "advisor_dashboard"
        elif role == User.Role.ESTUDIANTE:
            return_url = "student_dashboard"
        else:
            return_url = "dashboard_redirect"

        context.update(
            {
                "user_data": get_topbar_user_data(self.request.user),
                "return_dashboard_url": return_url,
                **dashboard_data,
            }
        )
        return context


class MissedAppointmentsDashboardView(LoginRequiredMixin, TemplateView):
    template_name = "cases/missed_appointments_dashboard.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ASESOR, User.Role.ESTUDIANTE]:
            messages.error(request, "No autorizado.")
            return redirect("dashboard_redirect")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dashboard_data = MissedAppointmentsDashboardService.get_dashboard_data(self.request.user)

        context.update(
            {
                "user_data": get_topbar_user_data(self.request.user),
                **dashboard_data,
            }
        )
        return context


class CaseAppointmentDetailView(LoginRequiredMixin, DetailView):
    model = CaseAppointment
    template_name = "cases/case_appointment_detail.html"
    context_object_name = "appointment"
    pk_url_kwarg = "appointment_id"

    def get_queryset(self):
        return CaseAppointment.objects.select_related(
            "case",
            "case__beneficiary",
            "case__assigned_student",
            "case__advisor",
            "created_by",
        )

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ASESOR, User.Role.ESTUDIANTE]:
            messages.error(request, "No autorizado.")
            return redirect("dashboard_redirect")
        return super().dispatch(request, *args, **kwargs)

    def get_object(self, queryset=None):
        appointment = super().get_object(queryset)
        if self.request.user.role == User.Role.ASESOR and appointment.case.advisor_id != self.request.user.id:
            raise PermissionDenied
        if self.request.user.role == User.Role.ESTUDIANTE and appointment.case.assigned_student_id != self.request.user.id:
            raise PermissionDenied
        return appointment

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["user_data"] = get_topbar_user_data(self.request.user)
        context["classification"] = self.object.missed_classification if self.object.is_missed else "Vigente"
        return context


class ReassignCaseView(LoginRequiredMixin, TemplateView):
    template_name = "cases/reassign_case.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != User.Role.SECRETARIA:
            messages.error(request, "No autorizado.")
            return redirect("dashboard_redirect")

        self.case = get_object_or_404(
            Case.objects.select_related(
                "beneficiary",
                "assigned_student",
                "advisor",
                "secretary",
            ),
            id=self.kwargs["case_id"],
        )
        return super().dispatch(request, *args, **kwargs)

    def get_students(self):
        active_statuses = [
            Case.CaseStatus.ASIGNADO,
            Case.CaseStatus.AUTOASIGNADO,
            Case.CaseStatus.EN_PROCESO,
            Case.CaseStatus.ESPERANDO_BENEFICIARIO,
            Case.CaseStatus.EN_REVISION,
        ]

        same_sala_filter = Q(
            assigned_cases__status__in=active_statuses,
            assigned_cases__category=self.case.category,
        )

        students = (
            User.objects.filter(
                role=User.Role.ESTUDIANTE,
                is_active=True,
            )
            .annotate(
                active_case_count=Count(
                    "assigned_cases",
                    filter=Q(assigned_cases__status__in=active_statuses),
                    distinct=True,
                ),
                sala_case_count=Count(
                    "assigned_cases",
                    filter=same_sala_filter,
                    distinct=True,
                ),
            )
            .exclude(id=self.case.assigned_student_id)
            .order_by("sala_case_count", "active_case_count", "first_name", "last_name")
        )
        return students

    def get_average_student_load(self):
        active_statuses = [
            Case.CaseStatus.ASIGNADO,
            Case.CaseStatus.AUTOASIGNADO,
            Case.CaseStatus.EN_PROCESO,
            Case.CaseStatus.ESPERANDO_BENEFICIARIO,
            Case.CaseStatus.EN_REVISION,
        ]

        counts = list(
            User.objects.filter(
                role=User.Role.ESTUDIANTE,
                is_active=True,
            )
            .annotate(
                active_case_count=Count(
                    "assigned_cases",
                    filter=Q(assigned_cases__status__in=active_statuses),
                    distinct=True,
                )
            )
            .values_list("active_case_count", flat=True)
        )
        if not counts:
            return 0
        return sum(counts) / len(counts)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["user_data"] = get_topbar_user_data(self.request.user)
        context["case"] = self.case
        context["students"] = self.get_students()
        context["average_student_load"] = self.get_average_student_load()
        return context

    def post(self, request, *args, **kwargs):
        new_student_id = request.POST.get("new_student")
        reason = request.POST.get("reason", "").strip()

        if not new_student_id:
            messages.error(request, "Debes seleccionar un nuevo estudiante.")
            return redirect("reassign_case", case_id=self.case.id)

        try:
            new_student = User.objects.get(
                id=new_student_id,
                role=User.Role.ESTUDIANTE,
                is_active=True,
            )
        except User.DoesNotExist:
            messages.error(request, "El estudiante seleccionado no es válido.")
            return redirect("reassign_case", case_id=self.case.id)

        if self.case.assigned_student == new_student:
            messages.error(
                request,
                "El caso ya está asignado a este estudiante.",
            )
            return redirect("reassign_case", case_id=self.case.id)

        if not self.get_students().filter(id=new_student.id).exists():
            messages.error(
                request,
                "El estudiante seleccionado no está disponible para reasignación.",
            )
            return redirect("reassign_case", case_id=self.case.id)

        with transaction.atomic():
            CaseAssignment.objects.create(
                case=self.case,
                assigned_student=new_student,
                assigned_advisor=request.user,
                assignment_type=CaseAssignment.AssignmentType.MANUAL,
                case_category=self.case.category or "",
                notes=reason or None,
            )

            self.case.assigned_student = new_student
            self.case.status = Case.CaseStatus.ASIGNADO
            self.case.current_stage = Case.CaseStage.ASSIGNMENT
            self.case.save(
                update_fields=[
                    "assigned_student",
                    "status",
                    "current_stage",
                    "updated_at",
                ]
            )

            content = f"Caso reasignado a {getattr(new_student, 'full_name', str(new_student))}."
            if reason:
                content += f" Motivo: {reason}"

            BitacoraEntry.objects.create(
                case=self.case,
                author=request.user,
                entry_type=BitacoraEntry.EntryType.ASIGNACION,
                content=content,
            )

        messages.success(request, "Caso reasignado exitosamente.")
        return redirect("case_detail", case_id=self.case.id)


class CaseBitacoraView(LoginRequiredMixin, TemplateView):
    template_name = "cases/case_bitacora.html"

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()

        self.case = get_object_or_404(
            Case.objects.select_related(
                "beneficiary",
                "assigned_student",
                "advisor",
                "secretary",
            ),
            id=self.kwargs["case_id"],
        )

        if not self.can_access_bitacora(request.user):
            messages.error(request, "No tienes permiso para acceder a la bitácora de este caso.")
            return redirect("dashboard_redirect")

        return super().dispatch(request, *args, **kwargs)

    def can_access_bitacora(self, user):
        if user.role == User.Role.SECRETARIA:
            return True

        allowed_user_ids = {
            self.case.assigned_student_id,
            self.case.advisor_id,
        }
        return user.id in allowed_user_ids

    def get_entries(self):
        return (
            BitacoraEntry.objects.filter(case=self.case)
            .select_related("author")
            .order_by("-created_at")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["user_data"] = get_topbar_user_data(self.request.user)
        context["case"] = self.case
        context["entries"] = self.get_entries()
        context["form"] = kwargs.get("form", BitacoraEntryForm(initial={"notify": True}))
        context["case_summary"] = build_case_bitacora_summary(self.case)
        return context

    def post(self, request, *args, **kwargs):
        form = BitacoraEntryForm(request.POST, request.FILES)

        if form.is_valid():
            entry = form.save(commit=False)
            entry.case = self.case
            entry.author = request.user
            legacy_entry_type = request.POST.get("entry_type")
            valid_entry_types = {choice[0] for choice in BitacoraEntry.EntryType.choices}
            entry.entry_type = (
                legacy_entry_type
                if legacy_entry_type in valid_entry_types
                else BitacoraEntry.EntryType.EVENTO
            )
            try:
                with transaction.atomic():
                    entry.save()

                    if entry.starts_new_term and entry.term_due_at:
                        CaseDeadline.objects.create(
                            case=self.case,
                            created_by=request.user,
                            title=f"Nuevo término: {entry.get_event_type_display()}",
                            description=entry.content,
                            due_date=entry.term_due_at,
                        )

                    files = form.cleaned_data.get("files", [])
                    for file in files:
                        document = BitacoraDocument.objects.create(
                            entry=entry,
                            file=file,
                            original_name=file.name,
                            file_size=file.size,
                            content_type=getattr(file, "content_type", None),
                        )
                        notify_advisor_document_uploaded(document)
            except DjangoValidationError as e:
                messages.error(request, e.message)
                return self.render_to_response(self.get_context_data(form=form))

            # Leer destinatarios de correo seleccionados en el checklist
            # Si no se seleccionó ninguno, email_recipients=None → envía a todos con email
            email_recipients = None
            if entry.notify:
                raw_emails = request.POST.getlist("email_recipients")
                parsed = {e.strip() for e in raw_emails if e.strip() and "@" in e}
                email_recipients = parsed if parsed else None

            # Notificaciones fuera de la transacción para no bloquear el guardado
            logger = logging.getLogger(__name__)
            try:
                notify_case_participants_bitacora_entry(
                    entry,
                    send_email=entry.notify,
                    email_recipients=email_recipients,
                )
            except Exception:
                logger.exception("Error al enviar notificaciones de bitácora para entrada %s", entry.id)

            messages.success(request, "Entrada agregada correctamente a la bitácora.")
            return redirect("case_bitacora", case_id=self.case.id)

        messages.error(request, "No se pudo guardar la entrada. Revisa el formulario.")
        return self.render_to_response(self.get_context_data(form=form))


class CaseDistributionListView(LoginRequiredMixin, TemplateView):
    """
    Vista para listar todos los casos SIN_ASIGNAR que pueden distribuirse (HU3).
    Acceso solo para Secretaría y Asesores.
    """
    template_name = "cases/case_distribution_list.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ASESOR]:
            messages.error(
                request,
                "No tienes permiso para acceder a esta funcionalidad.",
            )
            return redirect("dashboard_redirect")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        cases = (
            Case.objects.filter(status=Case.CaseStatus.SIN_ASIGNAR)
            .select_related("beneficiary", "assigned_student", "advisor", "secretary")
            .order_by("-created_at")
        )

        cases_with_status = []
        complete_count = 0
        type_choices_by_category = get_category_type_choices()
        for case in cases:
            completion_check = CaseCompletionValidator.is_complete(case)
            documents_info = CaseDocumentService.get_required_documents_for_display(case)
            
            # Construir información de documentos para el template
            doc_status = {
                'DOCUMENTO': {
                    'exists': documents_info.get('DOCUMENTO', {}).get('exists', False),
                    'valid': documents_info.get('DOCUMENTO', {}).get('valid', False),
                },
                'RECIBO_SERVICIOS': {
                    'exists': documents_info.get('RECIBO_SERVICIOS', {}).get('exists', False),
                    'valid': documents_info.get('RECIBO_SERVICIOS', {}).get('valid', False),
                },
                'FOTO': {
                    'exists': documents_info.get('FOTO', {}).get('exists', False),
                    'valid': documents_info.get('FOTO', {}).get('valid', False),
                },
            }
            
            cases_with_status.append(
                {
                    "case": case,
                    "is_complete": completion_check["is_complete"],
                    "completion_details": completion_check,
                    "documents": doc_status,
                    "type_choices": type_choices_by_category.get(case.category, {}),
                }
            )
            if completion_check["is_complete"]:
                complete_count += 1

        context.update(
            {
                "user_data": get_topbar_user_data(self.request.user),
                "cases": cases_with_status,
                "casos_completos_count": complete_count,
                "categories": Case.CaseCategory.choices,
                "students": User.objects.filter(
                    role=User.Role.ESTUDIANTE,
                    is_active=True,
                ).order_by("first_name", "last_name"),
            }
        )
        return context

    def post(self, request, *args, **kwargs):
        case_id = request.POST.get("case_id")
        action = request.POST.get("action")

        try:
            case = Case.objects.get(
                id=case_id,
                status=Case.CaseStatus.SIN_ASIGNAR,
            )
        except Case.DoesNotExist:
            return JsonResponse(
                {
                    "success": False,
                    "message": "El caso no existe o no está disponible.",
                },
                status=400,
            )

        completion_check = CaseCompletionValidator.is_complete(case)
        is_complete = completion_check["is_complete"]

        if action == "assign" and is_complete:
            return self._handle_assignment_json(request, case)
        elif action == "review" and not is_complete:
            return self._handle_send_to_review_json(request, case)
        else:
            return JsonResponse(
                {
                    "success": False,
                    "message": "Acción no válida para este caso.",
                },
                status=400,
            )

    def _handle_assignment_json(self, request, case):
        category = request.POST.get("category")
        case_type = request.POST.get("case_type")
        student_id = request.POST.get("assigned_student")

        if not all([category, case_type, student_id]):
            return JsonResponse(
                {
                    "success": False,
                    "message": "Debes completar todos los campos requeridos.",
                },
                status=400,
            )

        try:
            student = User.objects.get(
                id=student_id,
                role=User.Role.ESTUDIANTE,
                is_active=True,
            )
        except User.DoesNotExist:
            return JsonResponse(
                {
                    "success": False,
                    "message": "El estudiante seleccionado no es válido.",
                },
                status=400,
            )

        success, message, assignment = CaseDistributionService.assign_case_manually(
            case=case,
            student=student,
            category=category,
            case_type=case_type,
            assigned_by=request.user,
            notes="",
        )

        if success:
            return JsonResponse(
                {
                    "success": True,
                    "message": "¡Caso asignado exitosamente!",
                    "case_id": case.id,
                },
                status=200,
            )

        return JsonResponse(
            {
                "success": False,
                "message": message,
            },
            status=400,
        )

    def _handle_send_to_review_json(self, request, case):
        success, message = CaseDistributionService.send_to_review(
            case=case,
            sent_by=request.user,
            reason="Enviado a revisión desde el listado de distribución",
        )

        if success:
            return JsonResponse(
                {
                    "success": True,
                    "message": "Caso enviado a revisión exitosamente.",
                    "case_id": case.id,
                },
                status=200,
            )

        return JsonResponse(
            {
                "success": False,
                "message": message,
            },
            status=400,
        )


class CaseDistributionView(LoginRequiredMixin, TemplateView):
    """
    Vista para el reparto manual de casos (HU3).
    """
    template_name = "cases/case_distribution.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ASESOR]:
            messages.error(
                request,
                "No tienes permiso para acceder a esta funcionalidad.",
            )
            return redirect("dashboard_redirect")

        self.case = get_object_or_404(
            Case.objects.select_related(
                "beneficiary",
                "assigned_student",
                "advisor",
                "secretary",
            ),
            id=self.kwargs["case_id"],
        )
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        if self.case.status != Case.CaseStatus.SIN_ASIGNAR:
            context["error"] = (
                f"El caso no está disponible para reparto. "
                f"Estado: {self.case.get_status_display()}"
            )
            context["case"] = self.case
            context["user_data"] = get_topbar_user_data(self.request.user)
            return context

        completion_check = CaseCompletionValidator.is_complete(self.case)
        is_complete = completion_check["is_complete"]
        documents = CaseDocumentService.get_required_documents_for_display(self.case)
        form = kwargs.get("form") or CaseDistributionForm(case=self.case)

        context.update(
            {
                "user_data": get_topbar_user_data(self.request.user),
                "case": self.case,
                "is_complete": is_complete,
                "completion_details": completion_check,
                "documents": documents,
                "form": form,
                "button_text": "Asignar Caso" if is_complete else "Enviar a Revisión",
                "button_enabled": is_complete,
                "button_class": "btn-primary" if is_complete else "btn-warning",
            }
        )
        return context

    def post(self, request, *args, **kwargs):
        if self.case.status != Case.CaseStatus.SIN_ASIGNAR:
            messages.error(request, "El caso no está disponible para reparto.")
            return redirect("dashboard_redirect")

        completion_check = CaseCompletionValidator.is_complete(self.case)
        if completion_check["is_complete"]:
            return self._handle_assignment(request)
        return self._handle_send_to_review(request)

    def _handle_assignment(self, request):
        form = CaseDistributionForm(request.POST, case=self.case)

        if form.is_valid():
            success, message, assignment = CaseDistributionService.assign_case_manually(
                case=self.case,
                student=form.cleaned_data["assigned_student"],
                category=form.cleaned_data["category"],
                case_type=form.cleaned_data["case_type"],
                assigned_by=request.user,
                notes=form.cleaned_data.get("notes", "").strip(),
            )

            if success:
                messages.success(request, "¡Caso asignado exitosamente!")
                return redirect("case_detail", case_id=self.case.id)

            messages.error(request, message)
            return self.render_to_response(self.get_context_data(form=form))

        messages.error(request, "Por favor completa todos los campos requeridos.")
        return self.render_to_response(self.get_context_data(form=form))

    def _handle_send_to_review(self, request):
        reason = request.POST.get("reason", "").strip()

        success, message = CaseDistributionService.send_to_review(
            case=self.case,
            sent_by=request.user,
            reason=reason,
        )

        if success:
            messages.success(request, "Caso enviado a revisión exitosamente.")
            return redirect("case_detail", case_id=self.case.id)

        messages.error(request, message)
        return self.render_to_response(self.get_context_data())


class AssignedCasesCategoriesView(LoginRequiredMixin, TemplateView):
    template_name = "cases/assigned_cases_categories.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != User.Role.SECRETARIA:
            messages.error(
                request,
                "No tienes permiso para acceder a esta funcionalidad.",
            )
            return redirect("dashboard_redirect")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        assigned_cases = Case.objects.filter(
            status__in=[
                Case.CaseStatus.ASIGNADO,
                Case.CaseStatus.AUTOASIGNADO,
                Case.CaseStatus.EN_PROCESO,
                Case.CaseStatus.ESPERANDO_BENEFICIARIO,
                Case.CaseStatus.EN_REVISION,
            ]
        ).select_related("beneficiary", "assigned_student")

        categories = {}
        for case in assigned_cases:
            category = case.category if case.category else "Sin categoría"
            category_label = (
                case.get_category_display()
                if hasattr(case, "get_category_display")
                else category
            )

            if category not in categories:
                categories[category] = {
                    "label": category_label,
                    "count": 0,
                    "types": {},
                }

            categories[category]["count"] += 1

            case_type_short = case.case_type_specific if case.case_type_specific else "Sin tipo"
            case_type_label = get_type_label(case_type_short)

            if case_type_short not in categories[category]["types"]:
                categories[category]["types"][case_type_short] = {
                    "code": case_type_short,
                    "label": case_type_label,
                    "count": 0,
                }

            categories[category]["types"][case_type_short]["count"] += 1

        categories_list = [
            {
                "key": key,
                "label": value["label"],
                "count": value["count"],
                "types": sorted(value["types"].values(), key=lambda x: x["count"], reverse=True),
            }
            for key, value in categories.items()
        ]
        categories_list.sort(key=lambda x: x["count"], reverse=True)

        context["user_data"] = get_topbar_user_data(self.request.user)
        context["categories"] = categories_list
        context["total_assigned"] = assigned_cases.count()
        return context


class AssignedCasesAPIView(LoginRequiredMixin, View):
    def dispatch(self, request, *args, **kwargs):
        if request.user.role != User.Role.SECRETARIA:
            return JsonResponse({"error": "No autorizado"}, status=403)
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        category = request.GET.get("category", "")
        case_type = request.GET.get("type", "")

        cases_query = Case.objects.filter(
            status__in=[
                Case.CaseStatus.ASIGNADO,
                Case.CaseStatus.AUTOASIGNADO,
                Case.CaseStatus.EN_PROCESO,
                Case.CaseStatus.ESPERANDO_BENEFICIARIO,
                Case.CaseStatus.EN_REVISION,
            ]
        ).select_related("beneficiary", "assigned_student")

        if category and category != "Sin sala":
            cases_query = cases_query.filter(category=category)

        if case_type and case_type != "Sin trámite":
            cases_query = cases_query.filter(case_type_specific=case_type)

        cases_data = []
        for case in cases_query.order_by("-created_at"):
            cases_data.append(
                {
                    "id": case.id,
                    "case_number": case.case_number,
                    "beneficiary_name": case.beneficiary.full_name,
                    "student_name": (
                        case.assigned_student.full_name if case.assigned_student else "Sin asignar"
                    ),
                    "status": case.get_status_display(),
                    "category": (
                        case.get_category_display()
                        if hasattr(case, "get_category_display")
                        else case.category
                    ),
                    "type": (
                        get_type_label(case.case_type_specific)
                        if case.case_type_specific
                        else "Sin trámite"
                    ),
                    "description": (
                        case.description[:100] + "..."
                        if len(case.description) > 100
                        else case.description
                    ),
                }
            )

        return JsonResponse({"cases": cases_data})


class AssignedCasesFilteredView(LoginRequiredMixin, TemplateView):
    template_name = "cases/assigned_cases_filtered.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != User.Role.SECRETARIA:
            messages.error(
                request,
                "No tienes permiso para acceder a esta funcionalidad.",
            )
            return redirect("dashboard_redirect")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        category = self.request.GET.get("category", "")
        case_type_code = self.request.GET.get("type", "")
        category_label = (
            "Sin sala"
            if category == UNCATEGORIZED_CATEGORY_KEY
            else dict(Case.CaseCategory.choices).get(category, category)
        )
        case_type_label = (
            "Sin trámite"
            if case_type_code == UNTYPED_CASE_TYPE_KEY
            else get_type_label(case_type_code) if case_type_code else ""
        )

        cases_query = Case.objects.filter(
            status__in=[
                Case.CaseStatus.ASIGNADO,
                Case.CaseStatus.AUTOASIGNADO,
                Case.CaseStatus.EN_PROCESO,
                Case.CaseStatus.ESPERANDO_BENEFICIARIO,
                Case.CaseStatus.EN_REVISION,
            ]
        ).select_related("beneficiary", "assigned_student")

        if category and category != "Sin sala":
            cases_query = cases_query.filter(category=category)

        if case_type_code and case_type_code != "Sin trámite":
            cases_query = cases_query.filter(case_type_specific=case_type_code)

        context["user_data"] = get_topbar_user_data(self.request.user)
        context["cases"] = cases_query.order_by("-created_at")
        context["category"] = dict(Case.CaseCategory.choices).get(category, category)
        context["case_type"] = get_type_label(case_type_code) if case_type_code else ""
        context["cases_count"] = cases_query.count()
        return context


def panel_secretaria(request):
    casos_pendientes = Case.objects.filter(status=Case.CaseStatus.SIN_ASIGNAR)
    return render(request, "cases/panel_secretaria.html", {"casos": casos_pendientes})


def ejecutar_reparto(request, caso_id):
    if request.method == "POST":
        exito, mensaje = asignar_caso_automatico(caso_id, request.user)
        if exito:
            messages.success(request, mensaje)
        else:
            messages.error(request, mensaje)
    return redirect("panel_secretaria")


def calendario_seguimientos(request):
    seguimientos = BitacoraEntry.objects.all()
    return render(request, "cases/calendario.html", {"seguimientos": seguimientos})


class PendingCasesView(LoginRequiredMixin, ListView):
    model = Case
    template_name = "cases/pending_cases.html"
    context_object_name = "cases"
    paginate_by = 10

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ESTUDIANTE]:
            raise PermissionDenied
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return (
            Case.objects.filter(status=Case.CaseStatus.SIN_ASIGNAR)
            .select_related("beneficiary")
            .order_by("-created_at")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["total_pending"] = self.get_queryset().count()
        context["user_data"] = get_topbar_user_data(self.request.user)
        return context


class AssignedCaseDetailView(LoginRequiredMixin, DetailView):
    model = Case
    template_name = "cases/assigned_case_detail.html"
    context_object_name = "case"
    pk_url_kwarg = "case_id"

    def get_object(self, queryset=None):
        case = get_object_or_404(
            Case.objects.select_related(
                "beneficiary",
                "assigned_student",
                "advisor",
                "secretary",
            ),
            id=self.kwargs["case_id"],
        )
        user = self.request.user

        if user.role == User.Role.SECRETARIA:
            return case

        if user.role == User.Role.ESTUDIANTE:
            if case.assigned_student != user:
                raise PermissionDenied
            return case

        if user.role == User.Role.ASESOR:
            if case.advisor != user:
                raise PermissionDenied
            return case

        raise PermissionDenied

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["assignment"] = (
            CaseAssignment.objects.filter(case=self.object)
            .order_by("-assigned_at")
            .first()
        )
        context["binnacles"] = (
            self.object.bitacora_entries.select_related("author").order_by("-created_at")
        )
        context["case_documents"] = self.object.case_documents.all()
        context["user_data"] = get_topbar_user_data(self.request.user)
        return context

def reparto_automatico_view(request):
    """
    Vista para ejecutar el reparto automático equitativo de casos completos (HU3).
    """
    if request.method == "POST":
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ASESOR]:
            messages.error(request, "No tienes permiso para ejecutar el reparto automático.")
            return redirect("case_distribution_list")

        # Ejecutar reparto automático masivo
        success, message = CaseAutomaticDistributionService.ejecutar_reparto_automatico_masivo(
            executed_by=request.user
        )

        if success:
            messages.success(request, message)
            return redirect("case_distribution_list")
        else:
            messages.warning(request, message)
            return redirect("case_distribution_list")
    
    return redirect("case_distribution_list")


class CaseSearchAPIView(LoginRequiredMixin, View):
    """
    API endpoint para buscar casos por número.
    Solo usuarios autenticados pueden acceder.
    """
    
    def get(self, request):
        search_term = request.GET.get('q', '').strip()
        
        if not search_term or len(search_term) < 1:
            return JsonResponse({
                "results": [],
                "message": "Ingresa un número de caso para buscar"
            })
        
        # Buscar casos por número (case-insensitive)
        cases = Case.objects.filter(
            case_number__icontains=search_term
        ).select_related(
            'beneficiary',
            'assigned_student'
        ).order_by('-created_at')[:10]  # Máximo 10 resultados
        
        results = []
        for case in cases:
            results.append({
                'id': case.id,
                'case_number': case.case_number,
                'beneficiary_name': case.beneficiary.full_name if case.beneficiary else 'Sin beneficiario',
                'student_name': case.assigned_student.full_name if case.assigned_student else 'Sin asignar',
                'status': case.status,
                'status_display': case.get_status_display(),
            })
        
        return JsonResponse({
            "results": results,
            "count": len(results)
        })


class CaseSearchByIdAPIView(LoginRequiredMixin, View):
    """
    API endpoint para buscar un caso por ID exacto (5 dígitos numéricos).
    Busca TODOS los casos registrados (asignados y sin asignar).
    """
    
    def get(self, request):
        case_id = request.GET.get('id', '').strip()
        
        # Validar que sea exactamente 5 dígitos
        if not case_id or len(case_id) != 5 or not case_id.isdigit():
            return JsonResponse({
                "case": None,
                "error": "Ingresa un ID válido de 5 dígitos numéricos"
            }, status=400)
        
        # Buscar por ID exacto
        try:
            case = Case.objects.select_related(
                'beneficiary',
                'assigned_student'
            ).get(case_number=case_id)
            
            return JsonResponse({
                "case": {
                    'id': case.id,
                    'sequence_number': case.sequence_number,
                    'case_number': case.case_number,
                    'beneficiary_name': case.beneficiary.full_name if case.beneficiary else 'Sin beneficiario',
                    'student_name': case.assigned_student.full_name if case.assigned_student else 'Sin asignar',
                    'status': case.status,
                    'status_display': case.get_status_display(),
                    'created_at': case.created_at.strftime('%d/%m/%Y'),
                },
                "found": True
            })
        except Case.DoesNotExist:
            return JsonResponse({
                "case": None,
                "found": False,
                "error": f"No se encuentra un caso con ese ID"
            }, status=404)


class CaseSearchUnassignedByIdAPIView(LoginRequiredMixin, View):
    """
    API endpoint para buscar un caso SIN_ASIGNAR por ID exacto (5 dígitos numéricos).
    Usado para la barra de búsqueda en case_distribution_list (reparto de casos).
    """
    
    def get(self, request):
        case_id = request.GET.get('id', '').strip()
        
        # Validar que sea exactamente 5 dígitos
        if not case_id or len(case_id) != 5 or not case_id.isdigit():
            return JsonResponse({
                "case": None,
                "error": "Ingresa un ID válido de 5 dígitos numéricos"
            }, status=400)
        
        # Buscar por ID exacto entre casos SIN_ASIGNAR
        try:
            case = Case.objects.filter(
                status=Case.CaseStatus.SIN_ASIGNAR
            ).select_related(
                'beneficiary',
                'assigned_student'
            ).get(case_number=case_id)
            
            # Obtener información de completitud
            completion_check = CaseCompletionValidator.is_complete(case)
            
            return JsonResponse({
                "case": {
                    'id': case.id,
                    'sequence_number': case.sequence_number,
                    'case_number': case.case_number,
                    'beneficiary_name': case.beneficiary.full_name if case.beneficiary else 'Sin beneficiario',
                    'student_name': case.assigned_student.full_name if case.assigned_student else 'Sin asignar',
                    'status': case.status,
                    'status_display': case.get_status_display(),
                    'created_at': case.created_at.strftime('%d/%m/%Y'),
                    'is_complete': completion_check['is_complete'],
                },
                "found": True
            })
        except Case.DoesNotExist:
            return JsonResponse({
                "case": None,
                "found": False,
                "error": f"No se encuentra un caso con ese ID"
            }, status=404)


class CaseClosureAPIView(LoginRequiredMixin, View):
    """
    API endpoint para cerrar un caso y registrar el tipo de cierre.
    Usado cuando se cambia el estado a CERRADO en case_detail.
    Solo SECRETARIA y ASESOR pueden cerrar casos.
    """
    
    def post(self, request):
        # Validar permisos
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ASESOR]:
            return JsonResponse({
                "success": False,
                "error": "No tienes permiso para cerrar casos"
            }, status=403)
        
        case_id = request.POST.get('case_id')
        closure_type = request.POST.get('closure_type')
        closure_description = request.POST.get('closure_description', '').strip()
        closure_process_type = request.POST.get('closure_process_type') or Case.ClosureProcessType.JUDICIAL_PROCESS
        closure_reason = request.POST.get('closure_reason')

        reason_to_outcome = {
            Case.ClosureReason.DESISTIMIENTO_TACITO: Case.ClosureType.DISMISSED,
            Case.ClosureReason.DESISTIMIENTO_EXPRESO: Case.ClosureType.DISMISSED,
            Case.ClosureReason.GANADO: Case.ClosureType.FAVORABLE,
            Case.ClosureReason.PERDIDO: Case.ClosureType.NEGATIVE,
            Case.ClosureReason.INFRINGIO_TERMINOS: Case.ClosureType.NEGATIVE,
        }
        if closure_reason in reason_to_outcome and not closure_type:
            closure_type = reason_to_outcome[closure_reason]

        if not case_id or not closure_type:
            return JsonResponse({
                "success": False,
                "error": "Faltan parámetros requeridos"
            }, status=400)

        if closure_type not in [choice[0] for choice in Case.ClosureType.choices]:
            return JsonResponse({
                "success": False,
                "error": "Tipo de cierre inválido"
            }, status=400)

        if closure_process_type not in [choice[0] for choice in Case.ClosureProcessType.choices]:
            return JsonResponse({
                "success": False,
                "error": "Tipo de asunto inválido"
            }, status=400)

        if closure_reason and closure_reason not in [choice[0] for choice in Case.ClosureReason.choices]:
            return JsonResponse({
                "success": False,
                "error": "Motivo de cierre inválido"
            }, status=400)


        try:
            case = Case.objects.get(id=case_id)

            case.status = Case.CaseStatus.CERRADO
            case.closure_type = closure_type
            case.closure_description = closure_description
            case.closure_process_type = closure_process_type
            case.closure_reason = closure_reason or None
            case.current_stage = CaseStageManager.calculate_current_stage(case)
            case.save(update_fields=['status', 'closure_type', 'closure_description', 'closure_process_type', 'closure_reason', 'current_stage', 'updated_at'])

            # Crear entrada en bitácora
            closure_label = dict(Case.ClosureType.choices).get(closure_type, "Desconocido")
            process_label = dict(Case.ClosureProcessType.choices).get(closure_process_type, "Sin clasificar")
            reason_label = dict(Case.ClosureReason.choices).get(closure_reason, "Sin motivo registrado")
            BitacoraEntry.objects.create(
                case=case,
                author=request.user,
                entry_type=BitacoraEntry.EntryType.ACTUALIZACION,
                content=f"Caso cerrado: {process_label} - {closure_label}. Motivo: {reason_label}" + (f" Descripción: {closure_description}" if closure_description else "")
            )

            return JsonResponse({
                "success": True,
                "message": "Caso cerrado exitosamente"
            })
        except Case.DoesNotExist:
            return JsonResponse({
                "success": False,
                "error": "Caso no encontrado"
            }, status=404)
        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": str(e)
            }, status=500)


class ClosedCasesAPIView(LoginRequiredMixin, View):
    """
    API endpoint para obtener casos cerrados filtrados por tipo de cierre.
    Devuelve los casos en formato JSON para visualización en modal.
    """
    
    def get(self, request):
        # Validar permisos - acceso solo para secretaria y asesor
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ASESOR]:
            return JsonResponse({
                "success": False,
                "error": "No autorizado"
            }, status=403)
        
        closure_type = request.GET.get('closure_type', '').strip()
        process_type = request.GET.get('process_type', '').strip()
        
        # Validar tipo de cierre
        valid_types = [choice[0] for choice in Case.ClosureType.choices]
        if closure_type not in valid_types:
            return JsonResponse({
                "success": False,
                "error": "Tipo de cierre inválido"
            }, status=400)
        
        filters = {
            "status": Case.CaseStatus.CERRADO,
            "closure_type": closure_type,
        }
        if process_type:
            valid_process_types = [choice[0] for choice in Case.ClosureProcessType.choices]
            if process_type not in valid_process_types:
                return JsonResponse({
                    "success": False,
                    "error": "Tipo de asunto inválido"
                }, status=400)
            filters["closure_process_type"] = process_type

        # Obtener casos cerrados del tipo especificado
        cases = Case.objects.filter(
            **filters
        ).select_related(
            'beneficiary',
            'assigned_student',
            'advisor'
        ).order_by('-updated_at')
        
        cases_data = []
        for case in cases:
            cases_data.append({
                'id': case.id,
                'case_number': case.case_number,
                'beneficiary_name': case.beneficiary.full_name if case.beneficiary else 'Sin beneficiario',
                'student_name': case.assigned_student.full_name if case.assigned_student else 'Sin asignar',
                'advisor_name': case.advisor.full_name if case.advisor else 'Sin asesor',
                'closure_type_display': dict(Case.ClosureType.choices).get(closure_type, 'Desconocido'),
                'closure_description': case.closure_description or '',
                'closure_process_type_display': case.get_closure_process_type_display(),
                'closure_reason_display': case.get_closure_reason_display() if case.closure_reason else 'No registrado',
                'updated_at': case.updated_at.strftime('%d/%m/%Y'),
                'url': f'/cases/casos/{case.id}/'
            })
        
        return JsonResponse({
            "success": True,
            "cases": cases_data,
            "count": len(cases_data),
            "closure_type": closure_type,
            "closure_type_display": dict(Case.ClosureType.choices).get(closure_type, 'Desconocido')
        })


class CaseStatusChangeAPIView(LoginRequiredMixin, View):
    """
    API endpoint para cambiar el estado de un caso (sin cierre).
    Solo SECRETARIA y ASESOR pueden cambiar el estado.
    Sincroniza automáticamente las etapas del caso con el nuevo estado.
    """

    # Mapa: estado → etapa correspondiente
    STATUS_TO_STAGE = {
        Case.CaseStatus.SIN_ASIGNAR:       Case.CaseStage.UNASSIGNED,           # 0
        Case.CaseStatus.ASIGNADO:          Case.CaseStage.ASSIGNMENT,            # 1
        Case.CaseStatus.AUTOASIGNADO:      Case.CaseStage.INFORMATION_GATHERING, # 2
        Case.CaseStatus.EN_PROCESO:        Case.CaseStage.INFORMATION_GATHERING, # 2
        Case.CaseStatus.ESPERANDO_BENEFICIARIO: Case.CaseStage.ANALYSIS_DRAFTING,     # 3
        Case.CaseStatus.EN_REVISION:       Case.CaseStage.SUPERVISOR_REVIEW,     # 4
        Case.CaseStatus.DOCUMENTACION:     Case.CaseStage.UNASSIGNED,            # 0 (doc pendiente)
    }

    def post(self, request):
        # Validar permisos
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ASESOR]:
            return JsonResponse({
                "success": False,
                "error": "No tienes permiso para cambiar el estado del caso"
            }, status=403)

        case_id = request.POST.get('case_id')
        new_status = request.POST.get('status')

        if not case_id or not new_status:
            return JsonResponse({
                "success": False,
                "error": "Faltan parámetros requeridos"
            }, status=400)

        valid_statuses = [choice[0] for choice in Case.CaseStatus.choices]
        if new_status not in valid_statuses:
            return JsonResponse({
                "success": False,
                "error": "Estado inválido"
            }, status=400)

        try:
            case = Case.objects.get(id=case_id)

            # Determinar la nueva etapa según el estado
            new_stage = self.STATUS_TO_STAGE.get(new_status, Case.CaseStage.UNASSIGNED)

            # Aplicar cambios
            case.status = new_status
            case.current_stage = new_stage
            case.save(update_fields=['status', 'current_stage', 'updated_at'])

            # Crear entrada en bitácora
            status_label = dict(Case.CaseStatus.choices).get(new_status, "Desconocido")
            stage_label = dict(Case.CaseStage.choices).get(new_stage, "")
            BitacoraEntry.objects.create(
                case=case,
                author=request.user,
                entry_type=BitacoraEntry.EntryType.ACTUALIZACION,
                content=f"Estado del caso cambiado a: {status_label}. Etapa actual: {stage_label}"
            )

            mensaje = f'El caso "{case.case_number}" cambió a {status_label.lower()}'

            return JsonResponse({
                "success": True,
                "message": mensaje,
                "new_stage": new_stage,
                "new_stage_label": stage_label,
            })

        except Case.DoesNotExist:
            return JsonResponse({
                "success": False,
                "error": "Caso no encontrado"
            }, status=404)
        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": str(e)
            }, status=500)

def load_case_types(request):
    """
    Carga tipos de trámites según categoría(s) seleccionada(s).
    
    - 1 categoría: devuelve todos sus trámites
    - 2+ categorías: devuelve SOLO los trámites comunes (intersección)
    """
    # Obtener parámetro 'categories' (múltiples) o 'category' (legado single)
    categories_param = request.GET.get('categories') or request.GET.get('category', '')
    
    if not categories_param:
        return JsonResponse([], safe=False)
    
    # Dividir por comas si hay múltiples categorías
    categories_list = [c.strip() for c in categories_param.split(',') if c.strip()]
    
    if not categories_list:
        return JsonResponse([], safe=False)
    
    # Si es una sola categoría, devolver todos sus trámites
    if len(categories_list) == 1:
        category_id = categories_list[0]
        types_list = CaseDistributionForm.CATEGORY_TYPES.get(category_id, [])
        results = [{'id': type_id, 'name': type_name} for type_id, type_name in types_list]
        return JsonResponse(results, safe=False)
    
    # Si hay múltiples categorías: INTERSECCIÓN (solo trámites comunes)
    # Obtener IDs de trámites de la primera categoría como set
    first_category = categories_list[0]
    first_types = CaseDistributionForm.CATEGORY_TYPES.get(first_category, [])
    common_type_ids = {type_id for type_id, _ in first_types}
    
    # Hacer intersección con las otras categorías
    for category_id in categories_list[1:]:
        types_list = CaseDistributionForm.CATEGORY_TYPES.get(category_id, [])
        category_type_ids = {type_id for type_id, _ in types_list}
        common_type_ids = common_type_ids.intersection(category_type_ids)
    
    # Construir resultado manteniendo orden de la primera categoría
    results = []
    for type_id, type_name in first_types:
        if type_id in common_type_ids:
            results.append({'id': type_id, 'name': type_name})

    return JsonResponse(results, safe=False)


# ──────────────────────────────────────────────────────────────────────────────
# CALENDAR API VIEWS
# ──────────────────────────────────────────────────────────────────────────────

def _can_access_case_calendar(user, case):
    if user.role == User.Role.SECRETARIA:
        return True
    return user.id in {case.assigned_student_id, case.advisor_id}


def _parse_calendar_datetime(value):
    parsed = parse_datetime(value or "")
    if parsed and timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed)
    return parsed


def _calendar_event_end(start, end=None):
    return end or (start + timezone.timedelta(hours=1))


def _overlap_filter(start, end, start_field="start_datetime", end_field="end_datetime"):
    return Q(**{f"{start_field}__lt": end}) & (
        Q(**{f"{end_field}__gt": start}) |
        (Q(**{f"{end_field}__isnull": True}) & Q(**{f"{start_field}__gt": start - timezone.timedelta(hours=1)}))
    )


def _participant_cases(user):
    if not user or not user.is_authenticated:
        return Case.objects.none()
    if user.role == User.Role.ESTUDIANTE:
        return Case.objects.filter(assigned_student=user)
    if user.role == User.Role.ASESOR:
        return Case.objects.filter(advisor=user)
    return Case.objects.none()


def _case_calendar_participants(case):
    return [participant for participant in [case.assigned_student, case.advisor] if participant]


def _has_calendar_conflict(user, start, end, exclude_case=None):
    if not user or not start or not end:
        return False

    cases = _participant_cases(user)
    if exclude_case:
        cases = cases.exclude(id=exclude_case.id)

    if CaseCalendarEvent.objects.filter(case__in=cases).filter(
        _overlap_filter(start, end)
    ).exists():
        return True

    if AdvisorAvailabilitySlot.objects.filter(is_booked=True).filter(
        Q(advisor=user) | Q(booked_by=user) | Q(case__assigned_student=user)
    ).filter(start_datetime__lt=end, end_datetime__gt=start).exists():
        return True

    if BitacoraEntry.objects.filter(case__in=cases, scheduled_for__gte=start, scheduled_for__lt=end).exists():
        return True

    if CaseAppointment.objects.filter(case__in=cases, scheduled_for__gte=start, scheduled_for__lt=end).exclude(
        status=CaseAppointment.AppointmentStatus.CANCELLED
    ).exists():
        return True

    return False


def _calendar_conflict_message(users, start, end, exclude_case=None):
    for participant in users:
        if _has_calendar_conflict(participant, start, end, exclude_case=exclude_case):
            return f"{participant.full_name} ya tiene un evento o reunión en ese horario."
    return ""


def _busy_cases_for_calendar(user, case):
    if user.role == User.Role.ESTUDIANTE:
        return Case.objects.filter(assigned_student=user).exclude(id=case.id)
    if user.role == User.Role.ASESOR:
        return Case.objects.filter(advisor=user).exclude(id=case.id)
    if user.role == User.Role.SECRETARIA:
        filters = Q()
        if case.assigned_student_id:
            filters |= Q(assigned_student=case.assigned_student)
        if case.advisor_id:
            filters |= Q(advisor=case.advisor)
        if not filters:
            return Case.objects.none()
        return Case.objects.filter(filters).exclude(id=case.id)
    return Case.objects.none()


def _append_busy_blocks(events, user, case):
    busy_cases = _busy_cases_for_calendar(user, case).select_related("beneficiary")

    for ev in CaseCalendarEvent.objects.filter(case__in=busy_cases).select_related("case", "case__beneficiary"):
        events.append({
            "id": f"busy_cal_{ev.id}",
            "title": "Ocupado por otro caso",
            "start": ev.start_datetime.isoformat(),
            "end": _calendar_event_end(ev.start_datetime, ev.end_datetime).isoformat(),
            "display": "background",
            "backgroundColor": "#fee2e2",
            "borderColor": "#fecaca",
            "allDay": ev.is_all_day,
            "extendedProps": {
                "type": "busy",
                "description": f"Caso {ev.case.case_number} - {ev.case.beneficiary.full_name}",
                "can_delete": False,
            },
        })

    for slot in AdvisorAvailabilitySlot.objects.filter(
        is_booked=True,
    ).filter(
        Q(advisor=user) | Q(booked_by=user) | Q(case__in=busy_cases)
    ).exclude(case=case).select_related("case", "case__beneficiary", "booked_by"):
        events.append({
            "id": f"busy_slot_{slot.id}",
            "title": "Reunión en otro caso",
            "start": slot.start_datetime.isoformat(),
            "end": slot.end_datetime.isoformat(),
            "display": "background",
            "backgroundColor": "#fee2e2",
            "borderColor": "#fecaca",
            "allDay": False,
            "extendedProps": {
                "type": "busy",
                "description": f"Caso {slot.case.case_number} - {slot.case.beneficiary.full_name}",
                "can_delete": False,
            },
        })

    for entry in BitacoraEntry.objects.filter(
        case__in=busy_cases,
        scheduled_for__isnull=False,
    ).select_related("case", "case__beneficiary"):
        events.append({
            "id": f"busy_bit_{entry.id}",
            "title": "Evento en otro caso",
            "start": entry.scheduled_for.isoformat(),
            "end": (entry.scheduled_for + timezone.timedelta(hours=1)).isoformat(),
            "display": "background",
            "backgroundColor": "#fff1f2",
            "borderColor": "#fecdd3",
            "allDay": False,
            "extendedProps": {
                "type": "busy",
                "description": f"Caso {entry.case.case_number} - {entry.case.beneficiary.full_name}",
                "can_delete": False,
            },
        })


class CaseCalendarEventsAPIView(LoginRequiredMixin, View):
    def get(self, request, case_id):
        from django.utils import timezone as tz
        case = get_object_or_404(
            Case.objects.select_related("beneficiary", "assigned_student", "advisor"),
            id=case_id,
        )
        if not _can_access_case_calendar(request.user, case):
            return JsonResponse({"error": "Sin permiso"}, status=403)

        user = request.user
        events = []

        # 1. Manual calendar events
        for ev in case.calendar_events.select_related("created_by"):
            events.append({
                "id": f"cal_{ev.id}",
                "title": ev.title,
                "start": ev.start_datetime.isoformat(),
                "end": ev.end_datetime.isoformat() if ev.end_datetime else None,
                "color": ev.color,
                "allDay": ev.is_all_day,
                "extendedProps": {
                    "type": "calendar_event",
                    "event_type": ev.event_type,
                    "event_type_label": ev.get_event_type_display(),
                    "description": ev.description,
                    "teams_link": ev.teams_link,
                    "created_by": ev.created_by.full_name if ev.created_by else "",
                    "can_delete": (
                        user == ev.created_by or user.role == User.Role.SECRETARIA
                    ),
                },
            })

        # 2. Bitácora entries with scheduled date
        for entry in BitacoraEntry.objects.filter(case=case, scheduled_for__isnull=False).select_related("author"):
            events.append({
                "id": f"bit_{entry.id}",
                "title": entry.get_event_type_display(),
                "start": entry.scheduled_for.isoformat(),
                "color": "#8b5cf6",
                "allDay": False,
                "extendedProps": {
                    "type": "bitacora",
                    "description": entry.content[:200],
                    "created_by": entry.author.full_name if entry.author else "",
                    "can_delete": False,
                },
            })

        # 3. Deadlines
        for dl in case.deadlines.all():
            color = "#94a3b8" if dl.is_completed else "#f97316"
            events.append({
                "id": f"dl_{dl.id}",
                "title": f"⏰ {dl.title}",
                "start": dl.due_date.isoformat(),
                "color": color,
                "allDay": False,
                "extendedProps": {
                    "type": "deadline",
                    "description": dl.description,
                    "status": "Completado" if dl.is_completed else ("Vencido" if dl.is_overdue else "Pendiente"),
                    "can_delete": False,
                },
            })

        # 4. Appointments
        for apt in case.appointments.all():
            events.append({
                "id": f"apt_{apt.id}",
                "title": apt.title,
                "start": apt.scheduled_for.isoformat(),
                "color": "#0ea5e9",
                "allDay": False,
                "extendedProps": {
                    "type": "appointment",
                    "description": apt.notes,
                    "status": apt.get_status_display(),
                    "can_delete": False,
                },
            })

        # 5. Advisor availability slots
        if user.role in [User.Role.SECRETARIA, User.Role.ESTUDIANTE]:
            for slot in case.availability_slots.all().select_related("advisor", "booked_by"):
                events.append({
                    "id": f"avail_{slot.id}",
                    "title": f"Disponible — {slot.advisor.full_name}",
                    "start": slot.start_datetime.isoformat(),
                    "end": slot.end_datetime.isoformat(),
                    "color": "#22c55e",
                    "allDay": False,
                    "extendedProps": {
                        "type": "availability",
                        "advisor_name": slot.advisor.full_name,
                        "teams_link": slot.teams_link,
                        "slot_id": slot.id,
                        "is_booked": False,
                        "can_request": user.role == User.Role.ESTUDIANTE,
                        "can_delete": False,
                    },
                })
                events[-1]["title"] = (
                    f"Reunion reservada - {slot.advisor.full_name}" if slot.is_booked
                    else f"Disponible - {slot.advisor.full_name}"
                )
                events[-1]["color"] = "#64748b" if slot.is_booked else "#16a34a"
                events[-1]["extendedProps"]["is_booked"] = slot.is_booked
                events[-1]["extendedProps"]["booked_by"] = slot.booked_by.full_name if slot.booked_by else ""
                events[-1]["extendedProps"]["can_request"] = (
                    not slot.is_booked and user.role in [User.Role.ESTUDIANTE, User.Role.SECRETARIA]
                )
        elif user.role == User.Role.ASESOR:
            for slot in case.availability_slots.filter(advisor=user).select_related("booked_by"):
                color = "#64748b" if slot.is_booked else "#16a34a"
                title = (
                    f"Reservado — {slot.booked_by.full_name}" if slot.is_booked else "Disponible"
                )
                events.append({
                    "id": f"avail_{slot.id}",
                    "title": title,
                    "start": slot.start_datetime.isoformat(),
                    "end": slot.end_datetime.isoformat(),
                    "color": color,
                    "allDay": False,
                    "extendedProps": {
                        "type": "availability",
                        "advisor_name": user.full_name,
                        "teams_link": slot.teams_link,
                        "slot_id": slot.id,
                        "is_booked": slot.is_booked,
                        "booked_by": slot.booked_by.full_name if slot.booked_by else "",
                        "can_request": False,
                        "can_delete": not slot.is_booked,
                    },
                })

        _append_busy_blocks(events, user, case)
        return JsonResponse(events, safe=False)


class CaseCalendarEventCreateAPIView(LoginRequiredMixin, View):
    def post(self, request, case_id):
        case = get_object_or_404(Case, id=case_id)
        user = request.user

        if not _can_access_case_calendar(user, case):
            return JsonResponse({"error": "Sin permiso"}, status=403)
        if user.role == User.Role.ESTUDIANTE:
            return JsonResponse({"error": "Los estudiantes no pueden crear eventos"}, status=403)

        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "JSON inválido"}, status=400)

        title = data.get("title", "").strip()
        event_type = data.get("event_type", CaseCalendarEvent.EventType.OTRO)
        description = data.get("description", "")
        start_raw = data.get("start_datetime")
        end_raw = data.get("end_datetime")
        is_all_day = bool(data.get("is_all_day", False))
        teams_link = data.get("teams_link", "")

        if not title or not start_raw:
            return JsonResponse({"error": "Título y fecha de inicio son requeridos"}, status=400)

        from django.utils.dateparse import parse_datetime
        from django.utils import timezone as tz

        start = parse_datetime(start_raw)
        if start and tz.is_naive(start):
            start = tz.make_aware(start)

        end = None
        if end_raw:
            end = parse_datetime(end_raw)
            if end and tz.is_naive(end):
                end = tz.make_aware(end)

        if not start:
            return JsonResponse({"error": "Formato de fecha inválido"}, status=400)

        if end and end <= start:
            return JsonResponse({"error": "El fin debe ser posterior al inicio"}, status=400)

        conflict = _calendar_conflict_message(
            _case_calendar_participants(case),
            start,
            _calendar_event_end(start, end),
        )
        if conflict:
            return JsonResponse({"error": conflict}, status=400)

        valid_types = {c[0] for c in CaseCalendarEvent.EventType.choices}
        if event_type not in valid_types:
            event_type = CaseCalendarEvent.EventType.OTRO

        ev = CaseCalendarEvent.objects.create(
            case=case,
            created_by=user,
            event_type=event_type,
            title=title,
            description=description,
            start_datetime=start,
            end_datetime=end,
            is_all_day=is_all_day,
            teams_link=teams_link,
        )

        return JsonResponse({
            "id": f"cal_{ev.id}",
            "title": ev.title,
            "start": ev.start_datetime.isoformat(),
            "end": ev.end_datetime.isoformat() if ev.end_datetime else None,
            "color": ev.color,
            "allDay": ev.is_all_day,
        }, status=201)


class CaseCalendarEventDeleteAPIView(LoginRequiredMixin, View):
    def delete(self, request, case_id, event_id):
        ev = get_object_or_404(CaseCalendarEvent, id=event_id, case_id=case_id)
        user = request.user
        if user != ev.created_by and user.role != User.Role.SECRETARIA:
            return JsonResponse({"error": "Sin permiso"}, status=403)
        ev.delete()
        return JsonResponse({"ok": True})


class AdvisorAvailabilityCreateAPIView(LoginRequiredMixin, View):
    def post(self, request, case_id):
        case = get_object_or_404(Case, id=case_id)
        user = request.user

        if user.role != User.Role.ASESOR:
            return JsonResponse({"error": "Solo asesores pueden agregar disponibilidad"}, status=403)
        if case.advisor_id and case.advisor_id != user.id:
            return JsonResponse({"error": "Solo el asesor asignado puede agregar disponibilidad a este caso"}, status=403)

        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "JSON inválido"}, status=400)

        start_raw = data.get("start_datetime")
        end_raw = data.get("end_datetime")
        teams_link = data.get("teams_link", "")

        if not start_raw or not end_raw:
            return JsonResponse({"error": "Inicio y fin son requeridos"}, status=400)

        from django.utils.dateparse import parse_datetime
        from django.utils import timezone as tz

        start = parse_datetime(start_raw)
        end = parse_datetime(end_raw)
        if start and tz.is_naive(start):
            start = tz.make_aware(start)
        if end and tz.is_naive(end):
            end = tz.make_aware(end)

        if not start or not end:
            return JsonResponse({"error": "Formato de fecha inválido"}, status=400)
        if end <= start:
            return JsonResponse({"error": "El fin debe ser posterior al inicio"}, status=400)

        if AdvisorAvailabilitySlot.objects.filter(
            advisor=user,
            start_datetime__lt=end,
            end_datetime__gt=start,
        ).exists() or _has_calendar_conflict(user, start, end):
            return JsonResponse({"error": "Ya tienes disponibilidad, reunión o evento en ese horario"}, status=400)

        slot = AdvisorAvailabilitySlot.objects.create(
            advisor=user,
            case=case,
            start_datetime=start,
            end_datetime=end,
            teams_link=teams_link,
        )

        return JsonResponse({
            "id": f"avail_{slot.id}",
            "title": "Disponible",
            "start": slot.start_datetime.isoformat(),
            "end": slot.end_datetime.isoformat(),
            "color": "#16a34a",
        }, status=201)


class AdvisorAvailabilityDeleteAPIView(LoginRequiredMixin, View):
    def delete(self, request, case_id, slot_id):
        slot = get_object_or_404(AdvisorAvailabilitySlot, id=slot_id, case_id=case_id)
        if request.user != slot.advisor:
            return JsonResponse({"error": "Sin permiso"}, status=403)
        if slot.is_booked:
            return JsonResponse({"error": "No se puede eliminar una franja ya reservada"}, status=400)
        slot.delete()
        return JsonResponse({"ok": True})


class SendMeetingInvitationAPIView(LoginRequiredMixin, View):
    def post(self, request, case_id, slot_id):
        from django.core.mail import EmailMultiAlternatives
        from django.conf import settings as conf_settings

        case = get_object_or_404(
            Case.objects.select_related("beneficiary", "assigned_student", "advisor"),
            id=case_id,
        )
        slot = get_object_or_404(
            AdvisorAvailabilitySlot.objects.select_related("advisor", "booked_by"),
            id=slot_id, case_id=case_id,
        )
        user = request.user

        if user.role not in [User.Role.ESTUDIANTE, User.Role.SECRETARIA]:
            return JsonResponse({"error": "Sin permiso"}, status=403)
        if slot.is_booked:
            return JsonResponse({"error": "Esta franja ya está reservada"}, status=400)

        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            data = {}

        if not _can_access_case_calendar(user, case):
            return JsonResponse({"error": "Sin permiso"}, status=403)
        if slot.end_datetime <= timezone.now():
            return JsonResponse({"error": "No se puede reservar una franja vencida"}, status=400)

        participants_to_check = [slot.advisor]
        if user.role == User.Role.ESTUDIANTE:
            participants_to_check.append(user)
        elif case.assigned_student:
            participants_to_check.append(case.assigned_student)

        conflict = _calendar_conflict_message(
            participants_to_check,
            slot.start_datetime,
            slot.end_datetime,
        )
        if conflict:
            return JsonResponse({"error": conflict}, status=400)

        message = data.get("message", "").strip()
        teams_result = None
        if not slot.teams_link:
            teams_result = create_teams_calendar_event(
                subject=f"Reunion - Caso {case.case_number}",
                body=(
                    f"<p>Reunion del Consultorio Juridico Icesi para el caso "
                    f"{case.case_number} - {case.beneficiary.full_name}.</p>"
                ),
                start_datetime=slot.start_datetime,
                end_datetime=slot.end_datetime,
                attendees=[
                    slot.advisor.email,
                    case.assigned_student.email if case.assigned_student else "",
                    user.email,
                ],
            )
            if teams_result.created:
                slot.teams_link = teams_result.join_url

        slot.is_booked = True
        slot.booked_by = user
        slot.save(update_fields=["is_booked", "booked_by", "teams_link"])

        start_str = slot.start_datetime.strftime("%d/%m/%Y %H:%M")
        end_str = slot.end_datetime.strftime("%H:%M")
        subject = f"Solicitud de reunión — Caso {case.case_number}"

        teams_row = (
            f'<tr style="background:#f0f1ff;"><td style="padding:10px 14px;font-weight:700;">Enlace Teams</td>'
            f'<td style="padding:10px 14px;"><a href="{slot.teams_link}">{slot.teams_link}</a></td></tr>'
            if slot.teams_link else ""
        )
        msg_row = (
            f'<tr><td style="padding:10px 14px;font-weight:700;">Mensaje</td>'
            f'<td style="padding:10px 14px;">{message}</td></tr>'
            if message else ""
        )

        html_body = f"""
<html><body style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:20px;">
  <div style="background:#5454e8;color:white;padding:20px;border-radius:8px 8px 0 0;">
    <h2 style="margin:0;">📅 Solicitud de Reunión</h2>
    <p style="margin:6px 0 0;opacity:.85;">Consultorio Jurídico ICESI</p>
  </div>
  <div style="border:1px solid #d9def8;border-top:none;border-radius:0 0 8px 8px;padding:24px;">
    <p><strong>{user.full_name}</strong> ha solicitado una reunión.</p>
    <table style="width:100%;border-collapse:collapse;margin:16px 0;">
      <tr style="background:#f0f1ff;">
        <td style="padding:10px 14px;font-weight:700;">Caso</td>
        <td style="padding:10px 14px;">{case.case_number} — {case.beneficiary.full_name}</td>
      </tr>
      <tr>
        <td style="padding:10px 14px;font-weight:700;">Fecha y hora</td>
        <td style="padding:10px 14px;">{start_str} – {end_str}</td>
      </tr>
      {teams_row}
      {msg_row}
    </table>
    <p style="color:#5b6475;font-size:.88rem;">Por favor confirme o comuníquese directamente con el solicitante.</p>
  </div>
</body></html>"""

        plain_body = (
            f"Solicitud de reunión — Caso {case.case_number}\n\n"
            f"{user.full_name} ha solicitado una reunión.\n"
            f"Caso: {case.case_number} — {case.beneficiary.full_name}\n"
            f"Fecha: {start_str} – {end_str}\n"
            + (f"Teams: {slot.teams_link}\n" if slot.teams_link else "")
            + (f"Mensaje: {message}\n" if message else "")
        )

        recipient_users = [slot.advisor]
        if case.assigned_student:
            recipient_users.append(case.assigned_student)
        if user not in recipient_users:
            recipient_users.append(user)
        recipients = sorted({recipient.email for recipient in recipient_users if recipient and recipient.email})

        for recipient in [recipient for recipient in recipient_users if recipient and recipient != user]:
            Notification.objects.create(
                recipient=recipient,
                case=case,
                notification_type=Notification.NotificationType.IMPORTANT_EVENT,
                title=f"Reunion agendada - Caso {case.case_number}",
                message=f"{user.full_name} agendo una reunion para el {start_str} - {end_str}.",
            )

        logger_cal = logging.getLogger(__name__)
        if recipients:
            try:
                email_obj = EmailMultiAlternatives(subject, plain_body, conf_settings.DEFAULT_FROM_EMAIL, recipients)
                email_obj.attach_alternative(html_body, "text/html")
                email_obj.send()
            except Exception:
                logger_cal.exception("Error al enviar invitación de reunión para slot %s", slot.id)

        return JsonResponse({"ok": True, "message": "Invitación enviada correctamente"})
