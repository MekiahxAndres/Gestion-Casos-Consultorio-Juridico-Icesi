from collections import Counter
from pathlib import Path
from zipfile import BadZipFile, ZipFile
import re
import xml.etree.ElementTree as ET

from django.utils import timezone

from .models import BitacoraDocument, BitacoraEntry, CaseDeadline


MAX_TEXT_PER_DOCUMENT = 3500
MAX_DOCUMENTS_TO_READ = 8


def _compact_text(value):
    text = re.sub(r"\s+", " ", value or "").strip()
    return text[:MAX_TEXT_PER_DOCUMENT]


def _read_pdf(path):
    try:
        from pypdf import PdfReader
    except Exception:
        return ""

    try:
        reader = PdfReader(path)
        pages = []
        for page in reader.pages[:5]:
            pages.append(page.extract_text() or "")
        return _compact_text(" ".join(pages))
    except Exception:
        return ""


def _read_docx(path):
    try:
        with ZipFile(path) as archive:
            xml = archive.read("word/document.xml")
        root = ET.fromstring(xml)
        namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        return _compact_text(" ".join(node.text or "" for node in root.findall(".//w:t", namespace)))
    except (KeyError, BadZipFile, ET.ParseError, OSError):
        return ""


def _read_xlsx(path):
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        values = []
        for sheet in workbook.worksheets[:3]:
            for row in sheet.iter_rows(max_row=20, values_only=True):
                values.extend(str(cell) for cell in row if cell not in (None, ""))
        return _compact_text(" ".join(values))
    except Exception:
        return ""


def _read_pptx(path):
    try:
        snippets = []
        with ZipFile(path) as archive:
            for name in archive.namelist():
                if not name.startswith("ppt/slides/slide") or not name.endswith(".xml"):
                    continue
                root = ET.fromstring(archive.read(name))
                snippets.extend(node.text or "" for node in root.iter() if node.text)
                if len(snippets) > 120:
                    break
        return _compact_text(" ".join(snippets))
    except (BadZipFile, ET.ParseError, OSError):
        return ""


def extract_document_text(document):
    file_path = getattr(document.file, "path", "")
    if not file_path or not Path(file_path).exists():
        return ""

    suffix = Path(document.original_name or document.file.name).suffix.lower()
    if suffix == ".pdf":
        return _read_pdf(file_path)
    if suffix == ".docx":
        return _read_docx(file_path)
    if suffix == ".xlsx":
        return _read_xlsx(file_path)
    if suffix == ".pptx":
        return _read_pptx(file_path)
    return ""


def _latest_entries_summary(entries):
    snippets = []
    for entry in entries[:4]:
        date_text = timezone.localtime(entry.created_at).strftime("%d/%m/%Y")
        snippets.append(f"{date_text}: {entry.get_event_type_display()} - {_compact_text(entry.content)[:180]}")
    return snippets


def build_case_bitacora_summary(case):
    entries = list(
        BitacoraEntry.objects.filter(case=case)
        .select_related("author")
        .prefetch_related("documents")
        .order_by("-created_at")
    )
    documents = list(
        BitacoraDocument.objects.filter(entry__case=case)
        .select_related("entry", "entry__author")
        .order_by("-uploaded_at")
    )
    deadlines = list(
        CaseDeadline.objects.filter(case=case, is_completed=False).order_by("due_date")
    )
    now = timezone.now()
    event_counter = Counter(entry.get_event_type_display() for entry in entries)

    readable_documents = []
    unread_documents = []
    for document in documents[:MAX_DOCUMENTS_TO_READ]:
        extracted = extract_document_text(document)
        item = {
            "name": document.original_name or Path(document.file.name).name,
            "uploaded_at": document.uploaded_at,
            "text": extracted,
        }
        if extracted:
            readable_documents.append(item)
        else:
            unread_documents.append(item)

    document_insights = [
        f"{item['name']}: {item['text'][:220]}"
        for item in readable_documents[:3]
    ]

    next_deadlines = [
        {
            "title": deadline.title,
            "due_date": deadline.due_date,
            "is_overdue": deadline.due_date < now,
        }
        for deadline in deadlines[:3]
    ]

    scheduled_events = [
        entry
        for entry in entries
        if entry.scheduled_for and entry.scheduled_for >= now
    ][:3]

    highlights = []
    if entries:
        highlights.append(f"Se han registrado {len(entries)} entradas de seguimiento.")
    else:
        highlights.append("Aun no hay entradas registradas en la bitacora.")

    if documents:
        highlights.append(f"Hay {len(documents)} documento(s) adjunto(s) asociados al caso.")
    else:
        highlights.append("No hay documentos adjuntos en la bitacora.")

    if next_deadlines:
        overdue_count = sum(1 for item in next_deadlines if item["is_overdue"])
        if overdue_count:
            highlights.append(f"Existen {overdue_count} termino(s) vencido(s) o pendiente(s) de atencion.")
        else:
            highlights.append("Los terminos proximos estan identificados para seguimiento.")

    return {
        "entry_count": len(entries),
        "document_count": len(documents),
        "readable_document_count": len(readable_documents),
        "unread_document_count": len(unread_documents),
        "event_types": event_counter.most_common(4),
        "latest_entries": _latest_entries_summary(entries),
        "document_insights": document_insights,
        "next_deadlines": next_deadlines,
        "scheduled_events": scheduled_events,
        "highlights": highlights,
    }
